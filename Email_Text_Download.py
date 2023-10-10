#!/usr/bin/env python
# coding: utf-8

# In[5]:


import win32com.client
import openpyxl
import datetime
from datetime import timedelta

# Initialize Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# Define the folder where your Outlook emails are located
inbox = outlook.GetDefaultFolder(6)  # 6 corresponds to the Inbox folder

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Set the headers for the Excel file
worksheet['A1'] = 'Subject'
worksheet['B1'] = 'Received Date'
worksheet['C1'] = 'Sender'
worksheet['D1'] = 'Message Text'

# Get today's date
today = datetime.date.today()- timedelta(1)

# Iterate through the emails in the Inbox folder
row = 2  # Start from the second row (after headers)
for mail in inbox.Items:
    received_date = mail.ReceivedTime.date()

    # Check if the email was received today
    if received_date == today:
        subject = mail.Subject
        sender = mail.SenderName
        message_text = mail.Body

        # Write the email details to the Excel worksheet
        worksheet.cell(row=row, column=1, value=subject)
        worksheet.cell(row=row, column=2, value=received_date)
        worksheet.cell(row=row, column=3, value=sender)
        worksheet.cell(row=row, column=4, value=message_text)

        row += 1

# Save the Excel workbook
workbook.save(fr"\\AZ99\Accounting\Private\JBusser\01 Python\Email_Download\Archive\outlook_messages_{today}.xlsx")

# Close Outlook
outlook.Application.Quit()


# In[7]:


import pandas as pd


# In[8]:


df = pd.read_excel(fr"\\AZ99\Accounting\Private\JBusser\01 Python\Email_Download\Archive\outlook_messages_{today}.xlsx")


# In[15]:


df


# In[16]:


df_empty = df.loc[df['Sender'] == "Brandy Peet"]


# In[17]:


if df_empty.empty:
    print("DataFrame is empty")
else:
    print("DataFrame is not empty")


# In[23]:


def find_cells_with_both_texts_exclude_not(dataframe, column_name, text1, text2, exclusion_word="not"):
    """
    Find cells in a specific column where both text1 and text2 exist, excluding cells with exclusion_word in a Pandas DataFrame.

    Args:
    - dataframe (pd.DataFrame): The DataFrame to search in.
    - column_name (str): The name of the column to search in.
    - text1 (str): The first text to search for.
    - text2 (str): The second text to search for.
    - exclusion_word (str): The word to exclude from cells (default is "not").

    Returns:
    - List of tuples (row_index, cell_value) where both text1 and text2 are found and exclusion_word is not present in the specified column.
    """
    results = []

    for index, cell_value in enumerate(dataframe[column_name]):
        cell_str = str(cell_value)
        if text1 in cell_str and text2 in cell_str and exclusion_word not in cell_str:
            results.append((index, cell_value))

    return results




# In[33]:


import openpyxl

def change_cell_value(file_path, sheet_name, cell_address, new_value):
    """
    Change the value of a specific cell in an Excel workbook.

    Args:
    - file_path (str): The path to the Excel workbook.
    - sheet_name (str): The name of the worksheet.
    - cell_address (str): The cell address (e.g., 'A1', 'B2', etc.).
    - new_value: The new value to set in the cell.

    Returns:
    - None
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Select the specified worksheet
        worksheet = workbook[sheet_name]

        # Update the cell value
        worksheet[cell_address] = new_value

        # Save the changes to the workbook
        workbook.save(file_path)

        print(f"Cell {cell_address} in '{sheet_name}' changed to {new_value}.")

    except Exception as e:
        print(f"An error occurred: {str(e)}")




# In[36]:


column_name = "Subject"
search_text1 = "Cost Reports"
search_text2 = "NPL"



matches = find_cells_with_both_texts_exclude_not(df, column_name, search_text1, search_text2)

if matches:
    excel_file_path = r"\\AZ99\Accounting\Private\JBusser\00 Automation Production\Close_was_completed.xlsx"
    sheet_name = 'Sheet1'
    cell_address = 'B2'
    new_value = 1

    change_cell_value(excel_file_path, sheet_name, cell_address, new_value)
    print("Matches found:")
    for match in matches:
        print(f"Row {match[0]}, Column '{column_name}': {match[1]}")
else:
    print(f"No matches found in column '{column_name}'.")


# In[ ]:


column_name = "Subject"
search_text1 = "Cost Reports"
search_text2 = "Linetec"



matches = find_cells_with_both_texts_exclude_not(df, column_name, search_text1, search_text2)

if matches:
    excel_file_path = r"\\AZ99\Accounting\Private\JBusser\00 Automation Production\Close_was_completed.xlsx"
    sheet_name = 'Sheet1'
    cell_address = 'B3'
    new_value = 1

    change_cell_value(excel_file_path, sheet_name, cell_address, new_value)
    print("Matches found:")
    for match in matches:
        print(f"Row {match[0]}, Column '{column_name}': {match[1]}")
else:
    print(f"No matches found in column '{column_name}'.")


# In[ ]:


column_name = "Subject"
search_text1 = "Cost Reports"
search_text2 = "Neuco"



matches = find_cells_with_both_texts_exclude_not(df, column_name, search_text1, search_text2)

if matches:
    excel_file_path = r"\\AZ99\Accounting\Private\JBusser\00 Automation Production\Close_was_completed.xlsx"
    sheet_name = 'Sheet1'
    cell_address = 'B4'
    new_value = 1

    change_cell_value(excel_file_path, sheet_name, cell_address, new_value)
    print("Matches found:")
    for match in matches:
        print(f"Row {match[0]}, Column '{column_name}': {match[1]}")
else:
    print(f"No matches found in column '{column_name}'.")


# In[ ]:


column_name = "Subject"
search_text1 = "Cost Reports"
search_text2 = "National Powerline"



matches = find_cells_with_both_texts_exclude_not(df, column_name, search_text1, search_text2)

if matches:
    excel_file_path = r"\\AZ99\Accounting\Private\JBusser\00 Automation Production\Close_was_completed.xlsx"
    sheet_name = 'Sheet1'
    cell_address = 'B5'
    new_value = 1

    change_cell_value(excel_file_path, sheet_name, cell_address, new_value)
    print("Matches found:")
    for match in matches:
        print(f"Row {match[0]}, Column '{column_name}': {match[1]}")
else:
    print(f"No matches found in column '{column_name}'.")


# In[ ]:


column_name = "Subject"
search_text1 = "Cost Reports"
search_text2 = "Canyon"



matches = find_cells_with_both_texts_exclude_not(df, column_name, search_text1, search_text2)

if matches:
    excel_file_path = r"\\AZ99\Accounting\Private\JBusser\00 Automation Production\Close_was_completed.xlsx"
    sheet_name = 'Sheet1'
    cell_address = 'B6'
    new_value = 1

    change_cell_value(excel_file_path, sheet_name, cell_address, new_value)
    print("Matches found:")
    for match in matches:
        print(f"Row {match[0]}, Column '{column_name}': {match[1]}")
else:
    print(f"No matches found in column '{column_name}'.")

df.to_csv("test.csv")